from openpyxl import Workbook, load_workbook
from time import sleep
import pandas as pd


##wb = load_workbook('LRC.xlsx')
##ws = wb.active
##print(ws)
##print(ws['A1'].value)
##
##ws = wb['Preus CatSalut']
##for i in range(1,25):
##    print(ws['A{}'.format(i)].value)
    
##df = pd.read_excel('Libro1.xlsx',engine='openpyxl')

def compara(a,b):
    long = min(len(a),len(b))
    c = 0
    for i in range(long):
        if a[i] == b[i]:
            c+=1
    per_lletra = 100*(c/long)
    vsep = v.split()
    j = 0
    for vpar in vsep:
        if vpar in par:
            j+=1
   
    

    
    return per
        
a = 'CREATININA Sèrum'
b = 'Creatina-cinasa MB-Plasma [urgències]'
def get_list():
    ll = []
    wb = load_workbook('LRC.xlsx')
    ws = wb['Preus CatSalut']
    for i in range(1,1143):
        ll.append(ws['A{}'.format(i)].value)
    print(ll[-1])
    return ll

def classifica(per):
    ll_act = get_list()
    wb = load_workbook('LRC.xlsx')
    ws = wb['Catàleg LRC']
    for i in range(1,501):
        v = ws['B{}'.format(i)].value
        r = ''
        if v in ll_act:
            r = v
        else:
            vsep = v.split()
            for par in ll_act:
                j = 0
                for vpar in vsep:
                    if vpar in par:
                        j+=1


def get_list1():
    ll = []
    wb = load_workbook('LRC.xlsx')
    ws = wb['Catàleg LRC']
    for i in range(1,3301):
        ll.append(ws['B{}'.format(i)].value)
    print(ll[-1])
    return ll

def par_repes(ll):
    dic_par = {}
    for par in ll:
        par_sep = par.split()
        for ps in par_sep:
            if ps in dic_par:
                dic_par[ps] += 1
            else:
                dic_par[ps] = 1
    return dic_par
                        
def ordena():
    d = par_repes(get_list1())
    ll = []
    for clau in d:
        ll.append((d[clau],clau))
    ll.sort(reverse=True)
    return ll


wb = Workbook()
ws = wb.active
ws.title = 'paraules freqüents'
ll = ordena()
for tup in ll:
    ws.append(list(tup))

wb.save('paraules_freqüents.xlsx')










            
            
            
            
        
